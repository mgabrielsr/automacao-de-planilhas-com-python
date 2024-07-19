from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# 1- Lê pasta de trabalho e planilha
wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatorio"]

# 2- Referências das linhas e colunas
min_colum = wb.active.min_column
max_colum = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# 3- Adicionando dados e categorias no gráfico
barchart = BarChart()

data = Reference(
    sheet,
    min_col=min_colum + 1,
    max_col=max_colum,
    min_row=min_row,
    max_row=max_row
    
)

categories = Reference(
    sheet,
    min_col=min_colum,
    max_col=min_colum,
    min_row=min_row + 1,
    max_row=max_row
    
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# 4 - Criando o gráfico
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricantes"
barchart.style = 2

# 5- Salvando o workbook
wb.save("data/barchart.xlsx")